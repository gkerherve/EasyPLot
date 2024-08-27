
## LIBRARY------------------------------------------------------------------------
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import matplotlib.font_manager as font_manager

from scipy.signal import find_peaks
from openpyxl import load_workbook
## -------------------------------------------------------------------------------

## VARIABLES----------------------------------------------------------------------
## Seaborn.... Layer zorder
FileName = 'Data.xlsx'
SheetName = 'Ti2p'
CoreName = r'Ti 2p'

font_type = 'Verdana'
font_size = 12

Y_MAX =1.6
Y_MIN =0.8

BE_correction = True

x_min , x_max, step = 453, 469, 5   # For Ti2p
## PEAK ARRAY is defined as follow
## [name of the column, Color used, Show the label, Name of the label, alpha value, plot the column or ignore]
## allowing to ignore or show the column help to keep a record of the settings used
Peaks_Array = [
    ['Ti2p3 4+', 'blue', True, 'Ti2$p_{3/2} ^{4+}$', 0.6, True],
    ['Ti2p3 3+', 'red',  False, 'Ti2$p_{3/2} ^{3+}$', 0.6,  True],
    ['Ti2p3 2+', 'green', True, 'Ti2$p_{3/2} ^{2+}$', 0.6, True]
]
##-------------------------------------

## EXTRA PEAK LABEL is defined as
## [name of the label, X position, Y position (1 is the intensity of the MAX peak),
## Color used, used of Line (if Yes write 'Line'), rotation, show or ignore the label]
ExtraLabels_Array = [
    ['Ti 2$p_{1/2}$', 463.8, 0.55,'grey', 'Line', 90, True]
]
##-------------------------------------


## INSET
Inset = [455.08, 454.28, 10, 0.5, 'k', True]

## -------------------------------------------------------------------------------

##LOAD DATA-----------------------------------------------------------------------
# Load data from Excel file with specific folder path (using openpyxl)
# For Windows:
wb = load_workbook(filename= FileName) #, read_only=True)

# Select the sheet containing the data
sheet = wb[SheetName]
# Set the 'Sheet' worksheet as the active sheet
wb.active = sheet
wb.save(FileName)
# Create a list of lists to store the data
data_values = []

# Iterate through rows and columns to extract data
for row in sheet.iter_rows(values_only=True):
    data_values.append(row)

# Create a pandas DataFrame from the extracted data
# Assuming the first row contains column names
data = pd.DataFrame(data_values[1:], columns=data_values[0])
##for i in range(14) :
##    df = data.drop(2)
##    data = df
# Filter out rows with 0 values in all columns
data_filtered = data[(data != 0).all(axis=1)]


wb.close()
##---------------------------------------------------------------------------------

##INSET----------------------------------------------------------------------------
if Inset[5] :
    # Find the indices where the 'BE' column is equal to the start and end values
    start_indices = data['BE'] == Inset[0]
    end_indices = data['BE'] == Inset[1]


    # Get the first index where the condition is True for the start and end indices
    start_index = start_indices.idxmax()
    end_index = end_indices.idxmax()

    # Create the inset of the 'RAW DATA' column
    data_INSET_X = data.loc[start_index:end_index, 'BE']
    data_INSET = data.loc[start_index:end_index, 'RAW DATA']

    data_INSET = data_INSET - data_INSET.iloc[-1]
    data_INSET = data_INSET * Inset[2] + data_filtered['RAW DATA'].max() * Inset[3]
#----------------------------------------------------------------------------------


#BE CORRECTION---------------------------------------------------------------------
if BE_correction:
    BEcc = 284.8-data.iloc[0,0]

    data_filtered['BE'] = data_filtered['BE'] + BEcc
    data['BE'] = data['BE'] + BEcc
#----------------------------------------------------------------------------------



# Display the first few rows of the data
print(data_filtered.head())

# Calculate the new Y-axis maximum limit (20% higher than the current maximum intensity)
y_max_0 = data_filtered['RAW DATA'].max()
y_max = data_filtered['RAW DATA'].max() * Y_MAX
y_min = data_filtered['RAW DATA'].min() * Y_MIN






# Peaks
for Peak in Peaks_Array:
    if Peak[5]:
        plt.fill_between(
            data_filtered['BE'],
            data_filtered['Backgnd.'],
            data_filtered[Peak[0]],
            color=Peak[1], alpha=Peak[4])
        if Peak[2]:
            # Find local maxima (peak centers)
            peak_indices = np.argmax(data_filtered[Peak[0]]-data_filtered['Backgnd.'])
            peak_centers = data_filtered['BE'].iloc[peak_indices]
            # Add dashed lines at peak centers
            plt.plot([peak_centers, peak_centers], [0, y_max_0*1.1] , linestyle='--', color=Peak[1], alpha=0.9)
            plt.text(peak_centers , y_max_0*1.15 , Peak[3], fontsize=font_size, color=Peak[1], rotation=90,ha="center", va="bottom")


# Extra Labels
for ExtraLabel in ExtraLabels_Array:
    if ExtraLabel[6]:
        if ExtraLabel[4] == 'Line':
            plt.plot([ExtraLabel[1], ExtraLabel[1]], [0, y_max_0*ExtraLabel[2]], linestyle='--', color=ExtraLabel[3], alpha=0.9)
        plt.text(ExtraLabel[1], y_max_0 *(ExtraLabel[2]+0.05), ExtraLabel[0], fontsize=font_size, color=ExtraLabel[3], rotation=ExtraLabel[5], ha="center",
                 va="bottom")




# Create a scatter plot for XPS data
plt.scatter(data['BE'],data['RAW DATA'], color='k',label='XPS Data', s=15)

# Create a BKG and Envelope
plt.plot(data_filtered['BE'], data_filtered['Backgnd.'], color='darkgrey', linestyle='--')
plt.plot(data_filtered['BE'], data_filtered['Envelope'], color='k')

#Create INSET
if Inset[5]:
    plt.scatter(data_INSET_X, data_INSET, color=Inset[4], label='Inset', s=15)


##PLOT SETTING-------------------------------------------------------------------------
# Get the current axes
ax = plt.gca()

# Set the default font family
plt.rcParams['font.family'] = font_type

# Set the font size of the x-axis labels
font_size_X = font_size
font_props = font_manager.FontProperties(size=font_size)
ax.set_xticklabels(ax.get_xticklabels(), fontproperties=font_props)



# Set the custom subtick locator for the x-axis
ax.xaxis.set_minor_locator(ticker.AutoMinorLocator(4))

# Add text at the top left of the plot
plt.text(0.02, 0.95, CoreName, fontsize=font_size+8, fontweight='bold',
         fontfamily=font_type, transform=ax.transAxes, va='top', ha='left')


plt.xlabel(r'Binding energy (eV)', fontsize=font_size)
plt.ylabel(r'Intensity (a.u.)', fontsize=font_size)
plt.grid(False)

# Reverse the X-axis
plt.gca().invert_xaxis()

# Hide Y-axis values
plt.yticks([])

# Set the Y-axis limits
plt.ylim(y_min, y_max)
#ax.set_ylim([y_min, y_max])
plt.xlim(xmax=x_min, xmin=x_max)
#ax.set_xlim([356.2, 339.])

# Set the x-axis tick locations and labels
x_range = np.arange(x_min, x_max + step, step)

plt.gca().set_xticks(x_range)
plt.gca().set_xticklabels(x_range)



# Set the linewidth of the plot border (spines)
border_linewidth = 1.5
for spine in ax.spines.values():
    spine.set_linewidth(border_linewidth)
## END PLOT SETTINGS---------------------------------------------------------------------

plt.show()
