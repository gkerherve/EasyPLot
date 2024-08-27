
## LIBRARY------------------------------------------------------------------------
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

from scipy.signal import find_peaks
from openpyxl import load_workbook
## -------------------------------------------------------------------------------

## VARIABLES----------------------------------------------------------------------
FileName = 'Square.xlsx'
CoreName = 'Ca 2$\mathit{p}$'
## -------------------------------------------------------------------------------

# Load data from Excel file with specific folder path (using openpyxl)
# For Windows:
wb = load_workbook(filename= FileName, read_only=True)

# Select the sheet containing the data
sheet = wb.active

# Create a list of lists to store the data
data_values = []

# Iterate through rows and columns to extract data
for row in sheet.iter_rows(values_only=True):
    data_values.append(row)

# Create a pandas DataFrame from the extracted data
# Assuming the first row contains column names
data = pd.DataFrame(data_values[1:], columns=data_values[0])

# Filter out rows with 0 values in all columns
data_filtered = data[(data != 0).all(axis=1)]


wb.close()

# Display the first few rows of the data
print(data_filtered.head())

# Calculate the new Y-axis maximum limit (20% higher than the current maximum intensity)
y_max = data_filtered['RAW DATA'].max() * 1.4
y_min = data_filtered['RAW DATA'].min() * 0.9

# Find local maxima (peak centers)
peak_indices = np.argmax(data_filtered['Ca2p3'])
peak_centers = data_filtered['BE'].iloc[peak_indices]


# Peaks
plt.fill_between(data_filtered['BE'], data_filtered['Backgnd.'], data_filtered['Ca2p3'], color='blue', alpha=0.6)
plt.fill_between(data_filtered['BE'], data_filtered['Backgnd.'], data_filtered['Ca2p1'], color='blue', alpha=0.4)

# Add dashed lines at peak centers
#plt.axvline(peak_centers, linestyle='--', color='gray', alpha=0.6)
plt.plot([peak_centers, peak_centers], [0, y_max/1.4*1.1] , linestyle='--', color='gray', alpha=0.6)

plt.text(peak_centers , y_max/1.4*1.15 , 'Ca 2$\mathit{p}_{3/2}$', rotation=90,ha="center", va="bottom")

# Create a scatter plot for XPS data
plt.scatter(data_filtered['BE'],data_filtered['RAW DATA'], color='k',label='XPS Data', s=15)

# Create a line plot
plt.plot(data_filtered['BE'], data_filtered['Backgnd.'], color='grey', linestyle='--')
plt.plot(data_filtered['BE'], data_filtered['Envelope'], color='k')


# Set the default font family
plt.rcParams['font.family'] = 'Calibri'

# Get the current axes
ax = plt.gca()

# Add text at the top left of the plot
plt.text(0.02, 0.95, CoreName, fontsize=20, fontweight='bold',
         fontfamily='calibri', transform=ax.transAxes, va='top', ha='left')

plt.xlabel('Binding energy / eV')
#plt.ylabel('Intensity / a.u.')
plt.grid(False)

# Reverse the X-axis
plt.gca().invert_xaxis()

# Hide Y-axis values
plt.yticks([])

# Set the Y-axis limits
plt.ylim(y_min, y_max)
plt.xlim(354,344)

# Set the linewidth of the plot border (spines)
border_linewidth = 1.5
for spine in ax.spines.values():
    spine.set_linewidth(border_linewidth)


plt.show()
