import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import matplotlib.font_manager as font_manager
import matplotlib.lines as mlines
import matplotlib.cm as cm
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PilImage
import os

# SAVE DATA-----------------------------------------------------------------------
def SAVE_DATA(File,
              SheetName,
              DPI):
    #Load the workbook
    wb = load_workbook(filename= File)

    # Select the sheet containing the data
    sheet = wb[SheetName]
    # Set the 'Sheet' worksheet as the active sheet
    wb.active = sheet
    ws = wb.active

    # Load the image file
    #print('DATA/'+File+'_EasyPLOT_Fig.svg')
    plt.savefig( File+'_EasyPLOT_Fig.svg')
    plt.savefig('EasyPLOT_Fig.png',dpi=DPI)
    img = Image('EasyPLOT_Fig.png')

    # Add the image to the worksheet
    ws.add_image(img, 'C3')

    # Save the Excel file
    wb.save(filename= File)
    wb.close()
# --------------------------------------------------------------------------------

# SAVE DATA-----------------------------------------------------------------------
def SAVE_DATA2(CoreName,
              DPI):

    # Remove all spaces from the CoreName
    safe_core_name = CoreName.replace(' ', '')

    # Define the paths for the SVG and PNG files
    svg_filename = os.path.join('DATA', f"{safe_core_name}.svg")
    png_filename = os.path.join('DATA', f"{safe_core_name}.png")

    # Remove existing files if they exist
    if os.path.exists(svg_filename):
        os.remove(svg_filename)
    if os.path.exists(png_filename):
        os.remove(png_filename)

    # Save the figure
    try:
        plt.savefig(svg_filename)
        plt.savefig(png_filename, dpi=DPI)
    except Exception as e:
        print(f"Error saving figure: {e}")
# --------------------------------------------------------------------------------

# LOAD DATA-----------------------------------------------------------------------
# Load data from Excel file with specific folder path (using openpyxl)
# For Windows:
def LOAD_DATA(File,
              SheetName,
              OffsetRow):

    wb = load_workbook(filename= File)

    # Select the sheet containing the data
    sheet = wb[SheetName]
    # Set the 'Sheet' worksheet as the active sheet
    wb.active = sheet
    #wb.save(File)
    # Create a list of lists to store the data
    data_values = []

    # Iterate through rows and columns to extract data
    for row in sheet.iter_rows(values_only=True):
        data_values.append(row)

    # Create a pandas DataFrame from the extracted data
    # Assuming the first row contains column names
    data = pd.DataFrame(data_values[OffsetRow+2:], columns=data_values[OffsetRow])

    # Filter out rows with 0 values in all columns
    data_filter = data[(data != 0).all(axis=1)]


    wb.close()

    return(data_filter,
           data)
# ---------------------------------------------------------------------------------


# INSET----------------------------------------------------------------------------
def INSET(Inset,
          data,
          data_filtered):
    # Find the indices where the 'BE' column is equal to the start and end values
    start_indices = data['Binding Energy (calibrated)'] == Inset[0]
    end_indices = data['Binding Energy (calibrated)'] == Inset[1]


    # Get the first index where the condition is True for the start and end indices
    start_index = start_indices.idxmax()
    end_index = end_indices.idxmax()

    # Create the inset of the 'RAW DATA' column
    data_INSET_X = data.loc[start_index:end_index, 'Binding Energy (calibrated)']
    data_INSET = data.loc[start_index:end_index, 'Raw Data (Normalised)']

    data_INSET = data_INSET - data_INSET.iloc[-1]
    data_INSET = data_INSET * Inset[2] + data_filtered['Raw Data (Normalised)'].max() * Inset[3]

    return(data_INSET_X,
           data_INSET)
# ---------------------------------------------------------------------------------

# BE CORRECTION--------------------------------------------------------------------
def BE_CORRECTION(data_filtered,
                  data):

    BEcc = 284.8-data.iloc[0,0]

    data_filtered['Binding Energy (calibrated)'] = data_filtered['Binding Energy (calibrated)'] + BEcc
    data['Binding Energy (calibrated)'] = data['Binding Energy (calibrated)'] + BEcc

    return(data_filtered,
           data)
# ---------------------------------------------------------------------------------

# EXTRA LABEL----------------------------------------------------------------------
def EXTRA_LABEL(Fig, ax, y_max_0, ExtraLabels_Array, font_size):
    # Extra Labels-----------------------------
    for ExtraLabel in ExtraLabels_Array:
        if ExtraLabel[6]:
            if ExtraLabel[4] == 'Line':
                plt.plot([ExtraLabel[1],
                         ExtraLabel[1]],
                         [0, y_max_0*ExtraLabel[2]],
                         linestyle='--',
                         color=ExtraLabel[3],
                         alpha=0.9)

            plt.text(ExtraLabel[1],
                     y_max_0 *(ExtraLabel[2]+0.05),
                     ExtraLabel[0],
                     fontsize=font_size,
                     color=ExtraLabel[3],
                     rotation=ExtraLabel[5],
                     ha="center",
                     va="bottom")
    # ----------------------------------------
# ---------------------------------------------------------------------------------

# OTHER LABEL----------------------------------------------------------------------
def OTHER_LABEL(Fig, ax, OtherLabel_Array, font_size):
    for OtherLabel in OtherLabel_Array:
        if OtherLabel[5]:
            plt.text(OtherLabel[1],
                     OtherLabel[2],
                     OtherLabel[0],
                     fontsize=font_size,
                     color=OtherLabel[3],
                     rotation=OtherLabel[4],
                     transform=ax.transAxes,
                     ha='left',
                     va='bottom')
# ---------------------------------------------------------------------------------

# PLOTTING FUNCTION----------------------------------------------------------------
def PLOT(Fig,
         ax,
         IsFitPlot,                         # Fitting plot or line plot
         IsOffset,                          # Offset or no Offset in Line plot
         File_Number,                       # File Number
         ColorPerFile,                      # Color of the data
         data_filtered,                     # Data without the 0 values
         data,                              # All data
         Peaks_Array,                       # Peaks definition
         ExtraLabels_Array,                 # Peaks Labels definition
         Inset,                             # Inset definition
         data_INSET_X,                      # Data_inset X value
         data_INSET,                        # Data_inset Y value
         y_max_previous,                    # Get the previous y_max value
         Y_MAX,                             # Y MAX calculated from the highest peak
         Y_MIN,                             # Y MIN calculated from the lowest peak
         Y_Offset,                          # Offset between each plot
         Offset,
         y_max,
         font_size):                        #

    # Color initialise
    # cmap = cm.get_cmap('Set1')
    cmap = cm.get_cmap('tab20c')


    # Calculate the new Y-axis maximum limit (20% higher than the current maximum intensity)
    y_max_0 = data_filtered['Raw Data (Normalised)'].max()

    tmp_Offset = (y_max_0 - data_filtered['Raw Data (Normalised)'].min()) * Y_Offset

    y_min = data_filtered['Raw Data (Normalised)'].min() - tmp_Offset * 1.5
    
    # Peaks-----------------------------------
    for Peak in Peaks_Array:
        if IsFitPlot:
            if Peak[7]:                                             # SHOW the peak or IGNORE
                plt.fill_between(
                    data_filtered['Binding Energy (calibrated)'],
                    data_filtered['Backgnd. (Normalised)']+Offset + y_max_previous,
                    data_filtered[Peak[0]] + Offset + y_max_previous,
                    color=cmap(Peak[1]), alpha=Peak[5], zorder=Peak[6])



                if Peak[3]:                                         # SHOW the label or IGNORE
                    # Find local maxima (peak centers)
                    peak_indices = np.argmax(data_filtered[Peak[0]]-data_filtered['Backgnd. (Normalised)'])
                    peak_centers = data_filtered['Binding Energy (calibrated)'].iloc[peak_indices]
                    # Add dashed lines at peak centers
                    plt.plot([peak_centers, peak_centers],
                             [y_min + Offset + y_max_previous, y_max_0 * 1.1 + Offset + y_max_previous],
                             linestyle='--',
                             color=cmap(Peak[1]),
                             alpha=0.9)

                if Peak[2]:
                    # Find local maxima (peak centers)
                    peak_indices = np.argmax(data_filtered[Peak[0]]-data_filtered['Backgnd. (Normalised)'])
                    peak_centers = data_filtered['Binding Energy (calibrated)'].iloc[peak_indices]
                    plt.text(peak_centers,
                             y_max_0*1.15 + Offset + y_max_previous,
                             Peak[4],
                             fontsize=font_size,
                             color=cmap(Peak[1]),
                             rotation=90,
                             ha="center",
                             va="bottom")
    # -----------------------------------------

    # Create a scatter plot for XPS data------
    if IsFitPlot:
        plt.scatter(data['Binding Energy (calibrated)'],
                    data['Raw Data (Normalised)']+Offset +y_max_previous,
                    color=ColorPerFile[File_Number],
                    label='XPS Data',
                    linewidths=1.5,
                    s=15,
                    zorder=6)
    else:
        if IsOffset:
            plt.plot(data['Binding Energy (calibrated)'],
                     data['Raw Data (Normalised)']+Offset +y_max_previous,
                     color=ColorPerFile[File_Number])
        else:
            plt.plot(data['Binding Energy (calibrated)'],
                     data['Raw Data (Normalised)'],
                     color=ColorPerFile[File_Number])
    # ----------------------------------------

    # Create a BKG plot-----------------------
    if IsFitPlot:
        plt.plot(data_filtered['Binding Energy (calibrated)'],
                 data_filtered['Backgnd. (Normalised)']+Offset +y_max_previous,
                 color='darkgrey',
                 linestyle='--')
    # -----------------------------------------

    # Create a Envelope plot------------------
    if IsFitPlot:
        plt.plot(data_filtered['Binding Energy (calibrated)'],
                 data_filtered['Envelope (Normalised)']+Offset +y_max_previous,
                 color=ColorPerFile[File_Number],
                 linewidth = 2)
    # -----------------------------------------

    # Create INSET-----------------------------
    if Inset[5]:                                    # SHOW inset or IGNORE
        plt.scatter(data_INSET_X,
                    data_INSET,
                    color=Inset[4],
                    label='Inset',
                    s=15)
    # -----------------------------------------

    # OFFSET-----------------------------------
    if Offset == 0 or File_Number == 0:
        Offset = (y_max_0-data_filtered['Raw Data (Normalised)'].min())*Y_Offset

        y_max = data_filtered['Raw Data (Normalised)'].max() * Y_MAX


    y_min = Offset * Y_MIN

    y_max_previous = y_max_previous + y_max_0 + Offset
    if not IsOffset and not IsFitPlot:
        y_max_previous = 0
    # ----------------------------------------

    return y_min, y_max, y_max_previous, Offset
# ---------------------------------------------------------------------------------

# PLOT SETTING --------------------------------------------------------------------
def PLOT_SETTING(Fig,
                 ax,
                 font_type,
                 font_size,
                 LegendCustom,
                 Fig_w,
                 Fig_h,
                 X_MIN,
                 X_MAX,
                 y_min,
                 y_max,
                 STEP,
                 CoreName):
    # PLOT SETTING-----------------------------

    # Frame Modification --------
    print(Fig_h / 2.54)
    Fig.set_size_inches(Fig_w / 2.54,
                        Fig_h / 2.54)  #


    # Adjust the spacing between the subplot and the figure edges
    plt.subplots_adjust(left=0.05,
                        right=0.95,
                        bottom=0.08,
                        top=0.99)



    # Set the default font family
    plt.rcParams['font.family'] = font_type

    # Set the font size of the x-axis labels
    font_size_X = font_size
    font_props = font_manager.FontProperties(size=font_size)
    ax.set_xticklabels(ax.get_xticklabels(), fontproperties=font_props)

    # Set the custom subtick locator for the x-axis
    ax.xaxis.set_minor_locator(ticker.AutoMinorLocator(4))

    # Add text at the top left of the plot
    plt.text(0.01,
             0.99,
             CoreName,
             fontsize=font_size + 4,
             fontweight='bold',
             fontfamily=font_type,
             transform=ax.transAxes,
             va='top',
             ha='left')

    plt.xlabel(r'Binding energy (eV)',
               fontsize=font_size)
    plt.ylabel(r'Normalised intensity (a.u.)',
               fontsize=font_size)
    plt.grid(False)

    # Reverse the X-axis
    plt.gca().invert_xaxis()

    # Hide Y-axis values
    plt.yticks([])

    # Set the Y-axis limits
    plt.ylim(y_min, y_max)

    plt.xlim(xmax=X_MIN, xmin=X_MAX)


    # Set the x-axis tick locations and labels
    x_range = np.arange(X_MIN, X_MAX + STEP, STEP)

    plt.gca().set_xticks(x_range)
    plt.gca().set_xticklabels(x_range)

    # Set the linewidth of the plot border (spines)
    border_linewidth = 2
    for spine in ax.spines.values():
        spine.set_linewidth(border_linewidth)

    # Create a custom legend
    plt.tight_layout()

    legend_handles = [mlines.Line2D([[], ],
                                    [],
                                    color=c,
                                    label=l,
                                    # font_size=17,
                                    linewidth=3) for c, l in zip(LegendCustom[0], LegendCustom[1])]
    if LegendCustom[2]:
        plt.legend(handles=legend_handles,
                   loc='upper right',
                   frameon=False,
                   fontsize=14).set_visible(True)
        # plt.legend(fontsize=14)

    # else:
    #     # plt.legend(handles=legend_handles,
    #     #            loc='upper right',
    #     #            frameon=False).set_visible(False)


    # plt.tight_layout()
    #  END PLOT SETTINGS-----------------------


# ---------------------------------------------------------------------------------